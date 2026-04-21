import json, re, sys
from pathlib import Path
import openpyxl

src = Path(sys.argv[1] if len(sys.argv)>1 else "data/source-media.php 2.xlsx")
root = Path(__file__).resolve().parents[1]
wb = openpyxl.load_workbook(src, data_only=True)
ws = wb[wb.sheetnames[0]]
headers = [ws.cell(1,c).value for c in range(1, ws.max_column+1)]

def clean(v):
    if v is None: return ''
    if isinstance(v, str): return re.sub(r'\s+', ' ', v).strip()
    return str(v).strip()

rows=[]
for r in range(2, ws.max_row+1):
    rec={headers[c-1]: clean(ws.cell(r,c).value) for c in range(1, ws.max_column+1)}
    if any(rec.values()):
        rec["id"]=f"NLEM-{r-1:04d}"
        rows.append(rec)

(root/"data"/"nlem.json").write_text(json.dumps(rows, ensure_ascii=False, indent=2), encoding="utf-8")
print(f"Built {len(rows)} rows to data/nlem.json")
